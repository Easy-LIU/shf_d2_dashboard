"""
SHF-D2 | Sentiment Data Pipeline
==================================
Fetches 12 consumer sentiment series from FRED, cleans the master Excel,
outputs sentiment_clean.csv ready for the dashboard.

Usage:
  1. Put this file in the same folder as Master_DataSet.xlsx
  2. pip install pandas requests openpyxl
  3. python pipeline_final.py
"""

import pandas as pd
import requests
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIG ────────────────────────────────────────────────────────────────────

FRED_API_KEY = "1263ef12d515fa2651a2a5bd9f62a576"
MASTER_PATH  = "Master_DataSet.xlsx"
CLEAN_CSV    = "sentiment_clean.csv"

# 12 FRED sentiment series  {series_id: (metric_name, unit, source)}
FRED_SERIES = {
    "UMCSENT":          ("Michigan ICS – Overall",                     "Index (1966 Q1 = 100)", "University of Michigan – Surveys of Consumers"),
    "UMCSENTCURRENT":   ("Michigan ICC – Current Economic Conditions",  "Index (1966 Q1 = 100)", "University of Michigan – Surveys of Consumers"),
    "UMCSENTEXPECT":    ("Michigan ICE – Consumer Expectations",        "Index (1966 Q1 = 100)", "University of Michigan – Surveys of Consumers"),
    "MICH":             ("Michigan – 1yr Inflation Expectations",       "Percent (%)",           "University of Michigan – Surveys of Consumers"),
    "CSCICP03USM665S":  ("OECD CCI – Amplitude Adjusted",              "Normalised (avg=100)",  "OECD via FRED"),
    "USACSCICP02STSAM": ("OECD CCI – Standardised",                    "Normalised (avg=100)",  "OECD via FRED"),
    "CBCCST01USM661S":  ("Conference Board – Consumer Confidence",      "Index (1985=100)",      "Conference Board via FRED"),
    "EXPINF1YR":        ("Cleveland Fed – 1yr Inflation Expectations",  "Percent (%)",           "Federal Reserve Bank of Cleveland via FRED"),
    "EXPINF10YR":       ("Cleveland Fed – 10yr Inflation Expectations", "Percent (%)",           "Federal Reserve Bank of Cleveland via FRED"),
    "T5YIE":            ("5yr Breakeven Inflation Rate",                "Percent (%)",           "Federal Reserve Bank of St. Louis"),
    "ICPHSA":           ("Michigan – Buying Conditions: Houses",        "Index",                 "University of Michigan – Surveys of Consumers"),
    "ICPSA":            ("Michigan – Buying Conditions: Cars",          "Index",                 "University of Michigan – Surveys of Consumers"),
}

KEEP_METRIC_NAMES = {label for label, unit, source in FRED_SERIES.values()}

# Old metric names in the master → map to standard names above
LEGACY_NAME_MAP = {
    "Index of Consumer Sentiment (ICS)":                                          "Michigan ICS – Overall",
    "OECD Composite Consumer Confidence Index – US":                              "OECD CCI – Amplitude Adjusted",
    "One-Year Stock Market Confidence – Institutional":                           "Yale – Institutional 1yr Stock Confidence",
    "One-Year Stock Market Confidence – Individual":                              "Yale – Individual 1yr Stock Confidence",
    "Consumer Opinion Surveys: Composite Consumer Confidence (USACSCICP02STSAM)":"OECD CCI – Standardised",
    "Composite Consumer Confidence Amplitude Adjusted (CSCICP03USM665S)":         "OECD CCI – Amplitude Adjusted",
    "1-Year Expected Inflation (Cleveland Fed Model)":                            "Cleveland Fed – 1yr Inflation Expectations",
    "10-Year Expected Inflation (Cleveland Fed Model)":                           "Cleveland Fed – 10yr Inflation Expectations",
}

# Row fill color per data owner
OWNER_FILL = {
    "Raafay":    "D9E1F2",
    "Jalen":     "E2EFDA",
    "Jefferson": "FFF2CC",
    "Yijie":     "FCE4D6",
}

PLAIN_FONT = Font(name="Calibri", size=11)
LINK_FONT  = Font(name="Calibri", size=11, color="0563C1", underline="single")
ALIGN      = Alignment(vertical="top", wrap_text=False)

# Junk metric name keywords to drop
JUNK_KEYWORDS = [
    "OECD Data Filter", "Observation Frequency", "Series Start",
    "Series End", "Series Last", "Normalisation Benchmark",
    "Long-Run Normalisation", "Seasonal Adjustment", "Unit of Measurement",
    "Statistically Significant", "Series Observation", "Total Monthly",
    "Copyright", "Citation", "Number of Series", "Expansion Signal",
    "Composite Business Confidence Amplitude Adjusted – Normalisation",
    "Composite Business Confidence Amplitude Adjusted – Latest",
]


# ── FRED FETCH ────────────────────────────────────────────────────────────────

def fetch_fred(series_id: str) -> pd.DataFrame:
    """Fetch a single FRED series → DataFrame with [time_period, value]."""
    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&api_key={FRED_API_KEY}&file_type=json"
    )
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    obs = r.json()["observations"]
    df = pd.DataFrame(obs)
    df = df[df["value"] != "."].copy()
    df["value"] = pd.to_numeric(df["value"])
    df["time_period"] = pd.to_datetime(df["date"]).dt.to_period("M").astype(str)
    return df[["time_period", "value"]].reset_index(drop=True)


def fetch_all_fred() -> pd.DataFrame:
    """Fetch all 12 FRED series and return a combined DataFrame."""
    frames = []
    for series_id, (metric_name, unit, source) in FRED_SERIES.items():
        print(f"  Fetching {series_id}...")
        try:
            df = fetch_fred(series_id)
            df["Metric Name"]   = metric_name
            df["Unit"]          = unit
            df["Source"]        = f"https://fred.stlouisfed.org/series/{series_id}"
            df["Source/Survey"] = source
            df["Data Owner"]    = "Yijie"
            df["Notes"]         = f"FRED series {series_id}"
            frames.append(df)
            print(f"    ✓ {len(df)} rows")
        except Exception as e:
            print(f"    ✗ Skipped ({e})")
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ── CLEAN MASTER ──────────────────────────────────────────────────────────────

def clean_master(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean the existing master file:
    1. Drop metadata/junk rows
    2. Drop non-numeric values
    3. Drop non yyyy-mm dates
    4. Rename legacy metric names to standard names
    5. Keep only sentiment metrics + teammate rows
    6. Unify Data Owner (SHF_D2 → Yijie)
    7. Deduplicate
    """
    original_len = len(df)

    # Normalise string columns
    df["Metric Name"] = df["Metric Name"].astype(str).fillna("")
    df["Time Period"]  = df["Time Period"].astype(str).fillna("")

    # 1. Drop junk metric names
    mask_junk = df["Metric Name"].apply(
        lambda x: any(kw in x for kw in JUNK_KEYWORDS)
    )
    df = df[~mask_junk].copy()
    print(f"  Removed {mask_junk.sum()} metadata rows")

    # 2. Drop non-numeric values
    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
    before = len(df)
    df = df.dropna(subset=["Value"])
    print(f"  Removed {before - len(df)} non-numeric value rows")

    # 3. Drop non yyyy-mm dates
    before = len(df)
    df = df[df["Time Period"].str.match(r"^\d{4}-\d{2}$")]
    print(f"  Removed {before - len(df)} bad date format rows")

    # 4. Rename legacy metric names
    df["Metric Name"] = df["Metric Name"].replace(LEGACY_NAME_MAP)

    # 5. Keep sentiment metrics + teammate rows
    ALL_KEEP    = KEEP_METRIC_NAMES | set(LEGACY_NAME_MAP.values())
    team_owners = {"Raafay", "Jalen", "Jefferson"}
    before = len(df)
    mask = df["Metric Name"].isin(ALL_KEEP) | df["Data Owner"].isin(team_owners)
    df = df[mask].copy()
    print(f"  Removed {before - len(df)} off-topic rows")

    # 6. Unify owner label
    df["Data Owner"] = df["Data Owner"].replace("SHF_D2", "Yijie")

    # 7. Deduplicate
    before = len(df)
    df = df.drop_duplicates(subset=["Metric Name", "Time Period"])
    print(f"  Removed {before - len(df)} duplicate rows")

    if "#" in df.columns:
        df = df.drop(columns=["#"])

    print(f"  → {len(df)} clean rows (was {original_len})")
    return df


# ── MERGE & SORT ──────────────────────────────────────────────────────────────

def build_final(existing_clean: pd.DataFrame, new_fred: pd.DataFrame) -> pd.DataFrame:
    """Merge cleaned master + fresh FRED data, deduplicate, sort by date."""
    new_fred = new_fred.rename(columns={"time_period": "Time Period", "value": "Value"})

    cols = ["Data Owner", "Source", "Metric Name", "Value",
            "Unit", "Time Period", "Source/Survey", "Notes"]

    combined = pd.concat([existing_clean[cols], new_fred[cols]], ignore_index=True)
    combined = combined.drop_duplicates(subset=["Metric Name", "Time Period"])
    combined = combined.sort_values(["Time Period", "Metric Name"]).reset_index(drop=True)
    combined.insert(0, "#", range(1, len(combined) + 1))
    return combined


# ── WRITE EXCEL ───────────────────────────────────────────────────────────────

def write_excel(df: pd.DataFrame, master_path: str) -> None:
    """Rewrite Monthly Extraction sheet with clean sorted data + formatting."""
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = master_path.replace(".xlsx", f"_backup_{ts}.xlsx")
    shutil.copy(master_path, backup)
    print(f"  Backup: {backup}")

    wb = load_workbook(master_path)
    ws = wb["Monthly Extraction"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill()

    cols = ["#", "Data Owner", "Source", "Metric Name", "Value",
            "Unit", "Time Period", "Source/Survey", "Notes"]

    for i, row_data in enumerate(df[cols].itertuples(index=False), start=2):
        owner    = row_data[1]
        hex_col  = OWNER_FILL.get(owner, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=hex_col)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.alignment = ALIGN
            is_url = col_idx == 3 and isinstance(value, str) and value.startswith("http")
            cell.font = LINK_FONT if is_url else PLAIN_FONT

    wb.save(master_path)
    print(f"  ✓ Excel saved ({len(df)} rows)")


# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  SHF-D2 Sentiment Pipeline")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 55)

    # 1. Load
    print("\n[1/4] Loading master...")
    existing = pd.read_excel(MASTER_PATH, sheet_name="Monthly Extraction")
    print(f"  {len(existing)} rows loaded")

    # 2. Clean
    print("\n[2/4] Cleaning existing data...")
    clean = clean_master(existing)

    # 3. Fetch FRED
    print("\n[3/4] Fetching from FRED...")
    new_fred = fetch_all_fred()
    print(f"  {len(new_fred)} rows fetched from FRED")

    # 4. Merge & write
    print("\n[4/4] Merging and writing...")
    final = build_final(clean, new_fred)
    write_excel(final, MASTER_PATH)

    # Export CSV for dashboard
    final.to_csv(CLEAN_CSV, index=False)
    print(f"  ✓ CSV saved: {CLEAN_CSV}")

    # Summary
    print("\n── Final dataset ─────────────────────────────────────")
    summary = final.groupby(["Data Owner", "Metric Name"])["Value"].count()
    print(summary.sort_values(ascending=False).to_string())
    print(f"\nTotal: {len(final)} rows | {final['Metric Name'].nunique()} metrics")
    print("\n" + "=" * 55)
    print("  Done. Use sentiment_clean.csv for the dashboard.")
    print("=" * 55)
